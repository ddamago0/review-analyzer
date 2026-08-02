from io import BytesIO

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
import openpyxl
from openpyxl.styles import Font, PatternFill
import logging

logger = logging.getLogger(__name__)

router = APIRouter()


class ExportRequest(BaseModel):
    """
    Request body for exporting analysis results to Excel.
    """
    data: dict = Field(
        ...,
        description="Full analysis response returned by /api/upload"
    )


@router.post("/export/xlsx")
def export_excel(request: ExportRequest):
    """
    Export the analysis results to an Excel workbook.

    The workbook contains:
    - Resumen: dataset and statistics summary
    - Palabras más frecuentes: top word frequency table
    - Reseñas: the processed reviews
    - Análisis de tokens: token optimization projection (if present)

    Args:
        request (ExportRequest): Analysis response data

    Returns:
        StreamingResponse: Excel file download
    """
    try:
        wb = openpyxl.Workbook()

        _write_summary(wb.active, request.data)
        _write_word_frequency(wb, request.data)
        _write_reviews(wb, request.data)
        _write_token_analysis(wb, request.data)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = "analisis_resenas.xlsx"
        return StreamingResponse(
            buffer,
            media_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            },
        )
    except Exception as e:
        logger.error(f"Error exporting Excel: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error al exportar el archivo Excel: {str(e)}"
        )


def _style_header(ws, row, max_col):
    """Apply a consistent header style to a worksheet row."""
    fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font


def _write_summary(ws, data):
    """Write the summary sheet with dataset and statistics."""
    ws.title = "Resumen"

    summary_rows = []

    dataset = data.get("data", {})
    statistics = data.get("statistics", {})

    summary_rows.append(("Archivos procesados", data.get("files_processed", 0)))
    summary_rows.append(("Columna de reseñas detectada", dataset.get("column_name", "")))
    summary_rows.append(("Reseñas cargadas (originales)", dataset.get("total_original", 0)))
    summary_rows.append(("Duplicados eliminados", dataset.get("duplicates_removed", 0)))
    summary_rows.append(("Reseñas vacías eliminadas", dataset.get("empty_removed", 0)))
    summary_rows.append(("Reseñas procesadas", dataset.get("total_clean", 0)))

    summary_rows.append(("", ""))

    label_map = {
        "total_reviews": "Total de reseñas",
        "average_length": "Promedio de caracteres por reseña",
        "shortest_review": "Longitud mínima de reseña (caracteres)",
        "longest_review": "Longitud máxima de reseña (caracteres)",
        "average_words": "Promedio de palabras por reseña",
    }
    for key, label in label_map.items():
        if key in statistics:
            summary_rows.append((label, statistics[key]))

    for row_index, (label, value) in enumerate(summary_rows, start=1):
        ws.cell(row=row_index, column=1, value=label)
        ws.cell(row=row_index, column=2, value=value)

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 30


def _write_word_frequency(wb, data):
    """Write the word frequency sheet."""
    ws = wb.create_sheet("Palabras más frecuentes")
    ws.cell(row=1, column=1, value="#")
    ws.cell(row=1, column=2, value="Palabra")
    ws.cell(row=1, column=3, value="Frecuencia")
    _style_header(ws, 1, 3)

    word_frequency = data.get("word_frequency", [])
    for index, item in enumerate(word_frequency, start=2):
        ws.cell(row=index, column=1, value=index - 1)
        ws.cell(row=index, column=2, value=item.get("word", ""))
        ws.cell(row=index, column=3, value=item.get("count", 0))

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 15


def _write_reviews(wb, data):
    """Write the processed reviews sheet."""
    ws = wb.create_sheet("Reseñas")
    ws.cell(row=1, column=1, value="#")
    ws.cell(row=1, column=2, value="Reseña")
    _style_header(ws, 1, 2)

    reviews = data.get("data", {}).get("reviews", [])
    for index, review in enumerate(reviews, start=2):
        ws.cell(row=index, column=1, value=index - 1)
        ws.cell(row=index, column=2, value=review)

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 100


def _write_token_analysis(wb, data):
    """Write the token analysis sheet if present."""
    token_analysis = data.get("token_analysis")
    if not token_analysis or not token_analysis.get("enabled"):
        return

    ws = wb.create_sheet("Análisis de tokens")

    rows = [
        ("Reseñas analizadas", token_analysis.get("reviews_analyzed", 0)),
        ("Muestra traducida", token_analysis.get("translation_sample_size", 0)),
        ("Tokens originales (español)", token_analysis.get("total_original_tokens", 0)),
        ("Tokens traducidos (inglés)", token_analysis.get("total_translated_tokens", 0)),
        ("Diferencia de tokens", token_analysis.get("token_difference", 0)),
        ("Diferencia (%)", token_analysis.get("token_difference_percent", 0)),
        ("", ""),
    ]

    projection = token_analysis.get("projection", {})
    projection_labels = [
        ("Reseñas por día", "reviews_per_day"),
        ("Días por mes", "days_per_month"),
        ("Tokens originales por reseña", "original_tokens_per_review"),
        ("Tokens traducidos por reseña", "translated_tokens_per_review"),
        ("Tokens mensuales (original)", "original_monthly_tokens"),
        ("Tokens mensuales (traducido)", "translated_monthly_tokens"),
        ("Costo mensual original (USD)", "original_monthly_cost_usd"),
        ("Costo mensual traducido (USD)", "translated_monthly_cost_usd"),
        ("Ahorro mensual (USD)", "monthly_savings_usd"),
        ("Ahorro (%)", "savings_percent"),
        ("Precio por millón de tokens (USD)", "price_per_million_input_tokens_usd"),
    ]
    for label, key in projection_labels:
        if key in projection:
            rows.append((label, projection[key]))

    for row_index, (label, value) in enumerate(rows, start=1):
        ws.cell(row=row_index, column=1, value=label)
        ws.cell(row=row_index, column=2, value=value)

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 30
